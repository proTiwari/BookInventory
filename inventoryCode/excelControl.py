import pandas as pd
from openpyxl import load_workbook

from inventoryCode.ImageUploader import ImageUploader


def excelControl(allErrorList):

    wb = load_workbook('C:/Users/stabc/OneDrive/Desktop/python/excel/testing_first_book_into_inventory.xlsx')
    sheet = wb.worksheets[0]
    v = pd.read_excel(io='C:/Users/stabc/OneDrive/Desktop/python/excel/testing_first_book_into_inventory.xlsx')
    row_count = sheet.max_row
    j = 0
    resultList = []
    forwardedList = []
    for i in range(j, row_count - 1):

        Title = str(v['Title'].iloc[i])
        Author = str(v['Author'].iloc[i])
        Description = str(v['Description'].iloc[i])
        Subtitle = str(v['Subtitle'].iloc[i])
        MRP = str(v['MRP'].iloc[i])
        ISBN13 = str(v['ISBN13'].iloc[i])
        ISBN10 = str(v['ISBN10'].iloc[i])

        yearOfPublish = str(v['yearOfPublish'].iloc[i])
        backCoverImageUrl = str(v['backCoverImageName'].iloc[i])
        frontCoverImageUrl = str(v['frontCoverImageName'].iloc[i])
        supportingImages = str(v['supportingImages'].iloc[i])
        numberOfPages = str(v['NumberOfPages'].iloc[i])
        ImageUploaderList = ImageUploader(frontCoverImageUrl, backCoverImageUrl, supportingImages)
        if len(ImageUploaderList) == 1:
            return print("ImageUploaderList error(possible reason: image does not exist in the folder)")
        else:
            supportingImages = ImageUploaderList[0]
            frontCoverImageUrl = ImageUploaderList[1]
            backCoverImageUrl = ImageUploaderList[2]
        Author = Author.split(',')

        dic = {'Title': Title,
               'Author': Author,
               'Description': Description,
               'Subtitle': Subtitle,
               'MRP': MRP,
               'yearOfPublish': yearOfPublish,
               'backCoverImageUrl': backCoverImageUrl,
               'frontCoverImageUrl': frontCoverImageUrl,
               'supportingImages': supportingImages,
               'numberOfPages': numberOfPages,
               'ISBN10': ISBN10,
               'ISBN13': ISBN13}
        resultList.append(dic)
        # lenth = len(dic.keys())

    pd_xl_file = pd.ExcelFile("C:/Users/stabc/OneDrive/Desktop/python/excel/testing_first_book_into_inventory.xlsx")
    p = pd_xl_file.parse("Sheet1")
    dimensions = p.shape
    if dimensions[1] == 12:
        forwardedList.append(resultList)
        forwardedList.append(row_count)
        return forwardedList
    else:
        allErrorList.append(f"their should be 12 numbers of columns and you have provided {dimensions[1]}")
        return allErrorList
